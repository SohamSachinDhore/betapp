"""Export Manager for RickyMama data export functionality"""

import csv
import os
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import json

# Add src to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from utils.logger import get_logger


class ExportManager:
    """Handles data export to CSV/Excel with advanced filtering and formatting"""
    
    def __init__(self, db_manager, config_manager=None):
        self.db_manager = db_manager
        self.config_manager = config_manager
        self.logger = get_logger(__name__)
        
        # Get export configuration
        self.export_config = self._get_export_config()
        
        # Ensure export directory exists
        self._ensure_export_directory()
    
    def _get_export_config(self) -> Dict[str, Any]:
        """Get export configuration from config manager or defaults"""
        if self.config_manager:
            try:
                return self.config_manager.get_export_config()
            except:
                pass
        
        # Default configuration
        return {
            'default_path': './exports/',
            'format_options': ['CSV', 'Excel'],
            'max_export_rows': 100000,
            'include_headers': True,
            'date_format': '%d-%m-%Y',
            'encoding': 'utf-8'
        }
    
    def _ensure_export_directory(self):
        """Ensure export directory exists"""
        try:
            export_path = Path(self.export_config['default_path'])
            export_path.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            self.logger.error(f"Failed to create export directory: {e}")
    
    def export_universal_log(self, filters: Optional[Dict] = None, format_type: str = 'CSV') -> Dict[str, Any]:
        """Export universal log data"""
        try:
            # Get data from database
            data = self._get_universal_log_data(filters)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"universal_log_{timestamp}.{format_type.lower()}"
            filepath = Path(self.export_config['default_path']) / filename
            
            # Export based on format
            if format_type.upper() == 'CSV':
                success = self._export_to_csv(data, filepath, 'universal_log')
            else:
                success = self._export_to_excel(data, filepath, 'universal_log')
            
            if success:
                return {
                    'success': True,
                    'file_path': str(filepath),
                    'records_exported': len(data),
                    'message': f'Universal log exported successfully to {filename}'
                }
            else:
                return {
                    'success': False,
                    'error': 'Export failed',
                    'message': 'Failed to export universal log'
                }
                
        except Exception as e:
            self.logger.error(f"Universal log export failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'message': f'Export failed: {str(e)}'
            }
    
    def export_customers(self, format_type: str = 'CSV') -> Dict[str, Any]:
        """Export customers data"""
        try:
            # Get customers data
            customers = self.db_manager.get_all_customers()
            
            # Add statistics for each customer
            enhanced_data = []
            for customer in customers:
                customer_stats = self.db_manager.get_customer_statistics(customer['id'])
                enhanced_customer = {
                    'ID': customer['id'],
                    'Name': customer['name'],
                    'Created Date': customer['created_at'],
                    'Total Entries': customer_stats.get('total_entries', 0),
                    'Total Value': customer_stats.get('total_value', 0),
                    'Last Activity': customer_stats.get('last_activity', 'Never'),
                    'Avg Entry Value': customer_stats.get('avg_entry_value', 0)
                }
                enhanced_data.append(enhanced_customer)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"customers_{timestamp}.{format_type.lower()}"
            filepath = Path(self.export_config['default_path']) / filename
            
            # Export
            if format_type.upper() == 'CSV':
                success = self._export_to_csv(enhanced_data, filepath, 'customers')
            else:
                success = self._export_to_excel(enhanced_data, filepath, 'customers')
            
            if success:
                return {
                    'success': True,
                    'file_path': str(filepath),
                    'records_exported': len(enhanced_data),
                    'message': f'Customers exported successfully to {filename}'
                }
            else:
                return {
                    'success': False,
                    'error': 'Export failed',
                    'message': 'Failed to export customers'
                }
                
        except Exception as e:
            self.logger.error(f"Customers export failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'message': f'Export failed: {str(e)}'
            }
    
    def export_pana_table(self, filters: Optional[Dict] = None, format_type: str = 'CSV') -> Dict[str, Any]:
        """Export pana table data"""
        try:
            # Get pana table data
            data = self._get_pana_table_data(filters)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filter_suffix = ""
            if filters:
                if filters.get('bazar'):
                    filter_suffix += f"_{filters['bazar']}"
                if filters.get('date'):
                    filter_suffix += f"_{filters['date'].strftime('%Y%m%d')}"
            
            filename = f"pana_table{filter_suffix}_{timestamp}.{format_type.lower()}"
            filepath = Path(self.export_config['default_path']) / filename
            
            # Export
            if format_type.upper() == 'CSV':
                success = self._export_to_csv(data, filepath, 'pana_table')
            else:
                success = self._export_to_excel(data, filepath, 'pana_table')
            
            if success:
                return {
                    'success': True,
                    'file_path': str(filepath),
                    'records_exported': len(data),
                    'message': f'Pana table exported successfully to {filename}'
                }
            else:
                return {
                    'success': False,
                    'error': 'Export failed',
                    'message': 'Failed to export pana table'
                }
                
        except Exception as e:
            self.logger.error(f"Pana table export failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'message': f'Export failed: {str(e)}'
            }
    
    def export_time_table(self, filters: Optional[Dict] = None, format_type: str = 'CSV') -> Dict[str, Any]:
        """Export time table data"""
        try:
            # Get time table data
            data = self._get_time_table_data(filters)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filter_suffix = ""
            if filters:
                if filters.get('customer'):
                    filter_suffix += f"_{filters['customer'].replace(' ', '_')}"
                if filters.get('bazar'):
                    filter_suffix += f"_{filters['bazar']}"
                if filters.get('date'):
                    filter_suffix += f"_{filters['date'].strftime('%Y%m%d')}"
            
            filename = f"time_table{filter_suffix}_{timestamp}.{format_type.lower()}"
            filepath = Path(self.export_config['default_path']) / filename
            
            # Export
            if format_type.upper() == 'CSV':
                success = self._export_to_csv(data, filepath, 'time_table')
            else:
                success = self._export_to_excel(data, filepath, 'time_table')
            
            if success:
                return {
                    'success': True,
                    'file_path': str(filepath),
                    'records_exported': len(data),
                    'message': f'Time table exported successfully to {filename}'
                }
            else:
                return {
                    'success': False,
                    'error': 'Export failed',
                    'message': 'Failed to export time table'
                }
                
        except Exception as e:
            self.logger.error(f"Time table export failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'message': f'Export failed: {str(e)}'
            }
    
    def export_summary_data(self, format_type: str = 'CSV') -> Dict[str, Any]:
        """Export summary statistics"""
        try:
            # Get summary data
            summary_stats = self.db_manager.get_summary_statistics()
            
            # Format summary data for export
            summary_data = [
                {'Metric': 'Total Customers', 'Value': summary_stats.get('total_customers', 0)},
                {'Metric': 'Total Entries', 'Value': summary_stats.get('total_entries', 0)},
                {'Metric': 'Total Value', 'Value': summary_stats.get('total_value', 0)},
                {'Metric': 'Average Entry Value', 'Value': summary_stats.get('avg_entry_value', 0)},
                {'Metric': 'Most Active Customer', 'Value': summary_stats.get('most_active_customer', 'N/A')},
                {'Metric': 'Most Active Bazar', 'Value': summary_stats.get('most_active_bazar', 'N/A')},
                {'Metric': 'Export Date', 'Value': datetime.now().strftime(self.export_config['date_format'])},
            ]
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"summary_data_{timestamp}.{format_type.lower()}"
            filepath = Path(self.export_config['default_path']) / filename
            
            # Export
            if format_type.upper() == 'CSV':
                success = self._export_to_csv(summary_data, filepath, 'summary')
            else:
                success = self._export_to_excel(summary_data, filepath, 'summary')
            
            if success:
                return {
                    'success': True,
                    'file_path': str(filepath),
                    'records_exported': len(summary_data),
                    'message': f'Summary data exported successfully to {filename}'
                }
            else:
                return {
                    'success': False,
                    'error': 'Export failed',
                    'message': 'Failed to export summary data'
                }
                
        except Exception as e:
            self.logger.error(f"Summary export failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'message': f'Export failed: {str(e)}'
            }
    
    def export_multiple_tables(self, tables: List[str], filters: Optional[Dict] = None, format_type: str = 'Excel') -> Dict[str, Any]:
        """Export multiple tables to a single file"""
        try:
            if format_type.upper() != 'EXCEL':
                return {
                    'success': False,
                    'error': 'Multiple table export only supported for Excel format',
                    'message': 'Please use Excel format for multiple table export'
                }
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"multi_table_export_{timestamp}.xlsx"
            filepath = Path(self.export_config['default_path']) / filename
            
            # Try to import openpyxl
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                return {
                    'success': False,
                    'error': 'openpyxl not installed',
                    'message': 'Excel export requires openpyxl: pip install openpyxl'
                }
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            total_records = 0
            
            # Export each table to a separate sheet
            for table_name in tables:
                if table_name == 'universal':
                    data = self._get_universal_log_data(filters)
                    sheet_name = 'Universal Log'
                elif table_name == 'customers':
                    data = self._get_customers_export_data()
                    sheet_name = 'Customers'
                elif table_name == 'pana':
                    data = self._get_pana_table_data(filters)
                    sheet_name = 'Pana Table'
                elif table_name == 'time':
                    data = self._get_time_table_data(filters)
                    sheet_name = 'Time Table'
                elif table_name == 'summary':
                    data = self._get_summary_export_data()
                    sheet_name = 'Summary'
                else:
                    continue
                
                if not data:
                    continue
                
                # Create sheet
                ws = wb.create_sheet(title=sheet_name)
                
                # Add headers
                if data:
                    headers = list(data[0].keys())
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                    
                    # Add data
                    for row, record in enumerate(data, 2):
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=row, column=col, value=record.get(header, ''))
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    total_records += len(data)
            
            # Save workbook
            wb.save(filepath)
            
            return {
                'success': True,
                'file_path': str(filepath),
                'records_exported': total_records,
                'tables_exported': len(tables),
                'message': f'Multiple tables exported successfully to {filename}'
            }
            
        except Exception as e:
            self.logger.error(f"Multiple table export failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'message': f'Export failed: {str(e)}'
            }
    
    def create_full_backup(self) -> Dict[str, Any]:
        """Create a complete backup of all data"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Create backup directory
            backup_dir = Path(self.export_config['default_path']) / f"backup_{timestamp}"
            backup_dir.mkdir(exist_ok=True)
            
            backup_results = []
            total_records = 0
            
            # Backup each table
            tables_to_backup = ['universal', 'customers', 'pana', 'time', 'summary']
            
            for table in tables_to_backup:
                try:
                    if table == 'universal':
                        result = self.export_universal_log(format_type='CSV')
                    elif table == 'customers':
                        result = self.export_customers(format_type='CSV')
                    elif table == 'pana':
                        result = self.export_pana_table(format_type='CSV')
                    elif table == 'time':
                        result = self.export_time_table(format_type='CSV')
                    elif table == 'summary':
                        result = self.export_summary_data(format_type='CSV')
                    
                    if result['success']:
                        # Move file to backup directory
                        source_file = Path(result['file_path'])
                        backup_file = backup_dir / f"{table}_{source_file.name}"
                        source_file.rename(backup_file)
                        
                        backup_results.append({
                            'table': table,
                            'file': backup_file.name,
                            'records': result['records_exported']
                        })
                        total_records += result['records_exported']
                    
                except Exception as e:
                    self.logger.error(f"Failed to backup {table}: {e}")
                    backup_results.append({
                        'table': table,
                        'error': str(e)
                    })
            
            # Create backup manifest
            manifest = {
                'backup_date': datetime.now().isoformat(),
                'backup_timestamp': timestamp,
                'total_records': total_records,
                'tables': backup_results,
                'export_config': self.export_config
            }
            
            manifest_file = backup_dir / 'backup_manifest.json'
            with open(manifest_file, 'w', encoding='utf-8') as f:
                json.dump(manifest, f, indent=2, default=str)
            
            return {
                'success': True,
                'backup_path': str(backup_dir),
                'total_records': total_records,
                'tables_backed_up': len([r for r in backup_results if 'error' not in r]),
                'message': f'Full backup created successfully in {backup_dir.name}'
            }
            
        except Exception as e:
            self.logger.error(f"Full backup failed: {e}")
            return {
                'success': False,
                'error': str(e),
                'message': f'Backup failed: {str(e)}'
            }
    
    def _export_to_csv(self, data: List[Dict], filepath: Path, table_type: str) -> bool:
        """Export data to CSV file"""
        try:
            if not data:
                return False
            
            with open(filepath, 'w', newline='', encoding=self.export_config['encoding']) as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=data[0].keys())
                
                if self.export_config['include_headers']:
                    writer.writeheader()
                
                for row in data:
                    # Format data for CSV
                    formatted_row = self._format_row_for_export(row, 'CSV')
                    writer.writerow(formatted_row)
            
            self.logger.info(f"Successfully exported {len(data)} records to {filepath}")
            return True
            
        except Exception as e:
            self.logger.error(f"CSV export failed: {e}")
            return False
    
    def _export_to_excel(self, data: List[Dict], filepath: Path, table_type: str) -> bool:
        """Export data to Excel file"""
        try:
            # Try to import openpyxl
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                self.logger.error("openpyxl not installed - falling back to CSV")
                csv_path = filepath.with_suffix('.csv')
                return self._export_to_csv(data, csv_path, table_type)
            
            if not data:
                return False
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = table_type.replace('_', ' ').title()
            
            # Add headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            for row, record in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    value = record.get(header, '')
                    # Format value for Excel
                    formatted_value = self._format_value_for_excel(value)
                    ws.cell(row=row, column=col, value=formatted_value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filepath)
            
            self.logger.info(f"Successfully exported {len(data)} records to {filepath}")
            return True
            
        except Exception as e:
            self.logger.error(f"Excel export failed: {e}")
            return False
    
    def _format_row_for_export(self, row: Dict, format_type: str) -> Dict:
        """Format a row for export"""
        formatted_row = {}
        for key, value in row.items():
            if isinstance(value, (datetime, date)):
                formatted_row[key] = value.strftime(self.export_config['date_format'])
            elif isinstance(value, (int, float)) and 'value' in key.lower():
                # Format currency values
                formatted_row[key] = f"₹{value:,.2f}" if format_type == 'CSV' else value
            else:
                formatted_row[key] = str(value) if value is not None else ''
        return formatted_row
    
    def _format_value_for_excel(self, value):
        """Format a value for Excel export"""
        if isinstance(value, (datetime, date)):
            return value.strftime(self.export_config['date_format'])
        elif value is None:
            return ''
        else:
            return value
    
    # Data retrieval methods
    def _get_universal_log_data(self, filters: Optional[Dict] = None) -> List[Dict]:
        """Get universal log data for export"""
        try:
            return self.db_manager.get_universal_log_entries(
                filters=filters,
                limit=self.export_config['max_export_rows']
            )
        except Exception as e:
            self.logger.error(f"Failed to get universal log data: {e}")
            return []
    
    def _get_customers_export_data(self) -> List[Dict]:
        """Get customers data for export"""
        try:
            customers = self.db_manager.get_all_customers()
            enhanced_data = []
            
            for customer in customers:
                customer_stats = self.db_manager.get_customer_statistics(customer['id'])
                enhanced_customer = {
                    'ID': customer['id'],
                    'Name': customer['name'],
                    'Created Date': customer['created_at'],
                    'Total Entries': customer_stats.get('total_entries', 0),
                    'Total Value': customer_stats.get('total_value', 0),
                    'Last Activity': customer_stats.get('last_activity', 'Never'),
                    'Average Entry Value': customer_stats.get('avg_entry_value', 0)
                }
                enhanced_data.append(enhanced_customer)
            
            return enhanced_data
        except Exception as e:
            self.logger.error(f"Failed to get customers data: {e}")
            return []
    
    def _get_pana_table_data(self, filters: Optional[Dict] = None) -> List[Dict]:
        """Get pana table data for export"""
        try:
            return self.db_manager.get_pana_table_data(
                bazar=filters.get('bazar') if filters else None,
                date=filters.get('date') if filters else None,
                limit=self.export_config['max_export_rows']
            )
        except Exception as e:
            self.logger.error(f"Failed to get pana table data: {e}")
            return []
    
    def _get_time_table_data(self, filters: Optional[Dict] = None) -> List[Dict]:
        """Get time table data for export"""
        try:
            return self.db_manager.get_time_table_data(
                customer=filters.get('customer') if filters else None,
                bazar=filters.get('bazar') if filters else None,
                date=filters.get('date') if filters else None,
                limit=self.export_config['max_export_rows']
            )
        except Exception as e:
            self.logger.error(f"Failed to get time table data: {e}")
            return []
    
    def _get_summary_export_data(self) -> List[Dict]:
        """Get summary data for export"""
        try:
            summary_stats = self.db_manager.get_summary_statistics()
            
            return [
                {'Metric': 'Total Customers', 'Value': summary_stats.get('total_customers', 0)},
                {'Metric': 'Total Entries', 'Value': summary_stats.get('total_entries', 0)},
                {'Metric': 'Total Value', 'Value': summary_stats.get('total_value', 0)},
                {'Metric': 'Average Entry Value', 'Value': summary_stats.get('avg_entry_value', 0)},
                {'Metric': 'Most Active Customer', 'Value': summary_stats.get('most_active_customer', 'N/A')},
                {'Metric': 'Most Active Bazar', 'Value': summary_stats.get('most_active_bazar', 'N/A')},
                {'Metric': 'Export Date', 'Value': datetime.now().strftime(self.export_config['date_format'])},
            ]
        except Exception as e:
            self.logger.error(f"Failed to get summary data: {e}")
            return []
    
    def get_export_statistics(self) -> Dict[str, Any]:
        """Get export directory statistics"""
        try:
            export_path = Path(self.export_config['default_path'])
            if not export_path.exists():
                return {'total_files': 0, 'total_size': 0, 'last_export': None}
            
            files = list(export_path.glob('*.*'))
            total_size = sum(f.stat().st_size for f in files if f.is_file())
            
            # Get most recent export
            last_export = None
            if files:
                most_recent = max(files, key=lambda f: f.stat().st_mtime)
                last_export = datetime.fromtimestamp(most_recent.stat().st_mtime)
            
            return {
                'total_files': len(files),
                'total_size': total_size,
                'total_size_mb': round(total_size / (1024 * 1024), 2),
                'last_export': last_export,
                'export_path': str(export_path)
            }
        except Exception as e:
            self.logger.error(f"Failed to get export statistics: {e}")
            return {'error': str(e)}


def create_export_manager(db_manager, config_manager=None) -> ExportManager:
    """Factory function to create export manager"""
    return ExportManager(db_manager, config_manager)